"""Profile and normalize the supplied KPP crop workbook.

Usage: python scripts/import/profile_and_export.py [input.xlsx]
"""
from __future__ import annotations

import hashlib
import json
import math
import re
import sys
import uuid
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = ROOT / "upload" / "1.2.1 พืช-อำเภอต่อปี-new.xlsx"
OUTPUT_JSON = ROOT / "apps" / "web" / "public" / "data" / "crop-annual.json"
PROFILE_JSON = ROOT / "data" / "schema" / "profile.json"
PROFILE_MD = ROOT / "docs" / "data-profiling.md"

CROP_SHEETS = {
    "1-1 ข้าวนาปรัง": ("rice_offseason", "ข้าวนาปรัง"),
    "1-2 ข้าวนาปี": ("rice_main", "ข้าวนาปี"),
    "1-3 ข้าวโพด รุ่น1": ("maize_1", "ข้าวโพดรุ่น 1"),
    "1-4 ข้าวโพด รุ่น2": ("maize_2", "ข้าวโพดรุ่น 2"),
    "1-5 มันสำปะหลัง": ("cassava", "มันสำปะหลัง"),
    "1-6 ปาล์มน้ำมัน": ("oil_palm", "ปาล์มน้ำมัน"),
    "1-7 ยางพารา": ("rubber", "ยางพารา"),
    "1-8 อ้อย": ("sugarcane", "อ้อย"),
    "1-9 กล้อวยไข่": ("banana_egg", "กล้วยไข่"),
}

DISTRICT_CODES = {
    "เมืองกำแพงเพชร": "6201", "ไทรงาม": "6202", "คลองลาน": "6203",
    "ขาณุวรลักษบุรี": "6204", "คลองขลุง": "6205", "พรานกระต่าย": "6206",
    "ลานกระบือ": "6207", "ทรายทองวัฒนา": "6208", "ปางศิลาทอง": "6209",
    "บึงสามัคคี": "6210", "โกสัมพีนคร": "6211",
}


def clean_text(value):
    if value is None:
        return None
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def find_col(headers, tests):
    for idx, header in enumerate(headers):
        text = clean_text(header) or ""
        if all(test in text for test in tests):
            return idx
    return None


def record_id(district_code, year_be, crop_id):
    key = f"kpp:{district_code}:{year_be}:{crop_id}"
    return str(uuid.uuid5(uuid.NAMESPACE_URL, key))


def main():
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    wb = load_workbook(source, read_only=True, data_only=True)
    records, issues = [], []
    sheet_counts = {}

    for sheet_name, (crop_id, crop_name) in CROP_SHEETS.items():
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [clean_text(v) for v in rows[0]]
        district_i = find_col(headers, ["อำเภอ"])
        year_i = find_col(headers, ["ปี"])
        planted_i = find_col(headers, ["เนื้อที่เพาะปลูก"])
        harvested_i = find_col(headers, ["เนื้อที่เก็บเกี่ยว"])
        production_i = find_col(headers, ["ผลผลิต", "ตัน"])
        reported_planted_i = next((i for i,h in enumerate(headers) if h and "ผลผลิต" in h and "ปลูก" in h and "ต่อไร่" in h), None)
        reported_harvested_i = next((i for i,h in enumerate(headers) if h and "ผลผลิต" in h and ("เก็บ" in h or "กรีด" in h) and "ต่อไร่" in h), None)
        count = 0

        for row_no, row in enumerate(rows[1:], 2):
            district = clean_text(row[district_i]) if district_i is not None else None
            if not district:
                continue
            year_value = number(row[year_i]) if year_i is not None else None
            year_be = int(year_value) if year_value is not None else None
            planted = number(row[planted_i]) if planted_i is not None else None
            harvested = number(row[harvested_i]) if harvested_i is not None else None
            production = number(row[production_i]) if production_i is not None else None
            district_code = DISTRICT_CODES.get(district)
            flags = []
            quality = "pass"
            if not district_code or year_be is None:
                quality, flags = "error", ["required_invalid"]
            for label, value in (("planted", planted), ("harvested", harvested), ("production", production)):
                if value is not None and value < 0:
                    quality, flags = "error", flags + [f"negative_{label}"]
            if planted is not None and harvested is not None and harvested > planted:
                quality, flags = "warning", flags + ["harvested_gt_planted"]
            calculated = (production * 1000 / harvested) if production is not None and harvested and harvested > 0 else None
            rec = {
                "record_id": record_id(district_code or district, year_be, crop_id),
                "province_code": "62", "province_name": "กำแพงเพชร",
                "district_code": district_code, "district_name": district,
                "year_be": year_be, "year_ce": year_be - 543 if year_be else None,
                "crop_id": crop_id, "crop_name": crop_name,
                "planted_area_rai": planted, "harvested_area_rai": harvested,
                "production_ton": production,
                "reported_yield_planted_kg_rai": number(row[reported_planted_i]) if reported_planted_i is not None else None,
                "reported_yield_harvested_kg_rai": number(row[reported_harvested_i]) if reported_harvested_i is not None else None,
                "calculated_yield_kg_rai": calculated,
                "source_sheet": sheet_name, "source_row": row_no,
                "data_status": "published" if quality != "error" else "draft",
                "quality_status": quality, "quality_notes": flags,
            }
            records.append(rec)
            if flags:
                issues.append({"record_id": rec["record_id"], "sheet": sheet_name, "row": row_no, "issues": flags})
            count += 1
        sheet_counts[sheet_name] = count

    keys = Counter((r["district_code"], r["year_be"], r["crop_id"]) for r in records)
    duplicate_keys = [k for k,v in keys.items() if v > 1]
    for rec in records:
        if (rec["district_code"], rec["year_be"], rec["crop_id"]) in duplicate_keys:
            rec["quality_status"] = "error"
            rec["quality_notes"].append("duplicate_key")
            rec["data_status"] = "draft"

    monthly_ws = wb["ทบก. นาปี รายเดือน"]
    monthly_rows = list(monthly_ws.iter_rows(min_row=2, values_only=True))
    monthly_all_core_null = sum(1 for r in monthly_rows if r[3] is None and r[4] is None)
    status_counts = Counter(r["quality_status"] for r in records)
    crops = sorted({r["crop_name"] for r in records})
    districts = sorted({r["district_name"] for r in records})
    years = sorted({r["year_be"] for r in records if r["year_be"]})
    checksum = hashlib.sha256(source.read_bytes()).hexdigest()
    profile = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_file": source.name, "source_sha256": checksum,
        "sheet_count": len(wb.sheetnames), "sheet_names": wb.sheetnames,
        "annual_record_count": len(records), "annual_sheet_counts": sheet_counts,
        "crop_count": len(crops), "crops": crops,
        "district_count": len(districts), "districts": districts,
        "year_min": min(years), "year_max": max(years),
        "quality_counts": dict(status_counts),
        "duplicate_business_keys": len(duplicate_keys),
        "monthly_record_count": len(monthly_rows),
        "monthly_all_core_null_count": monthly_all_core_null,
        "issue_count": len(issues),
    }
    payload = {"meta": profile, "records": records}
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    PROFILE_JSON.parent.mkdir(parents=True, exist_ok=True)
    PROFILE_MD.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    PROFILE_JSON.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    lines = [
        "# รายงาน Data Profiling",
        "",
        f"- แหล่งข้อมูล: `{source.name}`",
        f"- SHA-256: `{checksum}`",
        f"- จำนวนชีต: {profile['sheet_count']} ชีต",
        f"- ข้อมูลรายปีจากชีตรายพืช: {len(records):,} ระเบียน",
        f"- ครอบคลุม: {len(crops)} พืช, {len(districts)} อำเภอ, พ.ศ. {min(years)}–{max(years)}",
        f"- สถานะคุณภาพ: ผ่าน {status_counts.get('pass',0):,}, เตือน {status_counts.get('warning',0):,}, ผิดพลาด {status_counts.get('error',0):,}",
        f"- Business key ซ้ำ: {len(duplicate_keys):,}",
        f"- ข้อมูลรายเดือน: {len(monthly_rows):,} ระเบียน; ค่าครัวเรือนและพื้นที่ว่างพร้อมกัน {monthly_all_core_null:,} ระเบียน",
        "",
        "## จำนวนระเบียนรายชีต",
        "",
        "| ชีต | ระเบียน |",
        "|---|---:|",
        *[f"| {name} | {count:,} |" for name,count in sheet_counts.items()],
        "",
        "## กฎสำคัญที่ใช้",
        "",
        "- ผลผลิตเฉลี่ยรวม = ผลผลิตรวม × 1,000 ÷ พื้นที่เก็บเกี่ยวรวม",
        "- ค่า null คงเป็น null และไม่แปลงเป็น 0",
        "- ระเบียนที่พื้นที่เก็บเกี่ยวมากกว่าพื้นที่เพาะปลูกติดสถานะ warning",
        "- ข้อมูลรายเดือนที่ตัวเลขหลักว่างทั้งหมดถูกกันออกจากข้อมูลเผยแพร่",
    ]
    PROFILE_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(profile, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
